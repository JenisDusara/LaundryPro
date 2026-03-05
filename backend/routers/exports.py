from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from db.database import get_db
from models import Customer, LaundryEntry

router = APIRouter(prefix="/api/exports", tags=["exports"])


def style_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="1e3a8a")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


@router.get("/customers")
def export_customers(
    token: str | None = None,
    db: Session = Depends(get_db),
):
    customers = db.query(Customer).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = ["Name", "Phone", "Flat No", "Society", "Address"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for c in customers:
        ws.append([c.name, c.phone, c.flat_number or "", c.society_name or "", c.address or ""])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=customers.xlsx"})


@router.get("/entries")
def export_entries(
    month: int,
    year: int,
    token: str | None = None,
    db: Session = Depends(get_db),
):
    entries = (
        db.query(LaundryEntry)
        .filter(
            extract("month", LaundryEntry.entry_date) == month,
            extract("year", LaundryEntry.entry_date) == year,
        )
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Entries"

    headers = ["Date", "Customer", "Phone", "Flat", "Society", "Service", "Qty", "Rate (Rs.)", "Amount (Rs.)", "Item Status", "Entry Status"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for entry in entries:
        for item in entry.items:
            ws.append([
                str(entry.entry_date),
                entry.customer.name,
                entry.customer.phone,
                entry.customer.flat_number or "",
                entry.customer.society_name or "",
                item.service_name,
                item.quantity,
                float(item.price_per_unit),
                float(item.subtotal),
                item.item_status,
                entry.delivery_status,
            ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=entries_{month}_{year}.xlsx"})